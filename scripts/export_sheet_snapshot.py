"""Read-only снимок обеих Google-таблиц для миграции на SQLite.

См. `sqlite_migration.md` §VI.1 (Э0). Это разовый инструмент в `scripts/`,
не часть бота — изменений в `src/` этот этап не делает.

Смешанный режим чтения — прямое следствие находки §I.3: колонка `A`
листа «Тикеты» обязана читаться как `FORMATTED_VALUE` (иначе настоящие даты
приходят серийными числами и `_parse_ticket_row`-подобный парсер их роняет),
а всё остальное — как `UNFORMATTED_VALUE`, чтобы суммы и формулы не
округлялись до отображаемой строки.

Каждая строка внутри диапазона данных пишется в CSV, включая полностью
пустые: легаси-`SUMIF` в старых формулах связывает «Тикеты» и «Юзеры»
позиционно (по номеру строки), так что пропуск строки сдвинул бы это
сопоставление у всех характеризационных тестов Э1.

Диапазоны блока `DataBase` (Тикеты/Юзеры/Магазин/item database) и верхней
таблицы «БУСТЫ» переиспользуют `stalbot.infrastructure.sheets.layouts` —
они откалиброваны против живой таблицы и уже проверены `validate_layout`
на старте бота.

Диапазоны убежки (`SHELTER_*`/`PROFESSION_SHEETS` ниже) и нижней таблицы
«БУСТЫ» откалиброваны вручную по офлайн-выгрузкам `СКУПКА.xlsx`/`убежка.xlsx`
(владелец скачал их «Файл → Скачать → .xlsx» — та же живая таблица, без
доступа к Sheets API). Источник данных не важен для остального конвейера:
`XlsxRangeReader` реализует тот же протокол `RangeReader`, что и
`GspreadRangeReader`, поэтому все функции `export_*` работают одинаково
что от живой таблицы, что от локального `.xlsx`-снимка.
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
from collections.abc import Sequence
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Final, Protocol

import gspread
import openpyxl
from google.oauth2.service_account import Credentials
from openpyxl.worksheet.worksheet import Worksheet

from stalbot.infrastructure.sheets.a1 import (
    a1_range,
    column_index_to_letter,
    column_letter_to_index,
    parse_a1_range,
)
from stalbot.infrastructure.sheets.layouts import (
    DATA_START_ROW,
    DATABASE_BLOCKS,
    DATABASE_SHEET,
    DatabaseBlock,
)

logger = logging.getLogger(__name__)

_SCOPES: Final = ("https://www.googleapis.com/auth/spreadsheets.readonly",)
_UNFORMATTED: Final = "UNFORMATTED_VALUE"
_FORMATTED: Final = "FORMATTED_VALUE"
_FORMULA: Final = "FORMULA"

#: Верхняя граница сканирования при поиске последней заполненной строки
#: блока. С большим запасом относительно текущих объёмов (§I.1: max 659).
_MAX_SCAN_ROWS: Final = 5_000

Row = list[object]


class RangeReader(Protocol):
    """Абстракция над одним batched-чтением диапазона — для тестируемости.

    Реализация поверх `gspread` живёт в `GspreadRangeReader`; в юнит-тестах
    подставляется фейк со словарём `{(sheet, a1, render_option): rows}`.
    """

    def __call__(self, sheet: str, a1: str, value_render_option: str) -> list[Row]:
        """Прочитать один A1-диапазон одним рендер-опшеном."""
        ...


@dataclass(frozen=True, slots=True)
class GspreadRangeReader:
    """`RangeReader`, читающий один открытый `gspread.Spreadsheet`."""

    spreadsheet: gspread.Spreadsheet

    def __call__(self, sheet: str, a1: str, value_render_option: str) -> list[Row]:
        """Прочитать один A1-диапазон через `Spreadsheet.values_get`."""
        response = self.spreadsheet.values_get(
            a1, params={"valueRenderOption": value_render_option}
        )
        values: list[list[object]] = response.get("values", [])
        return values


@dataclass(slots=True)
class XlsxRangeReader:
    """`RangeReader` поверх локального `.xlsx`-снимка (`openpyxl`).

    Держит две открытые книги: `_values` (`data_only=True`, кэшированные
    вычисленные значения — как `UNFORMATTED_VALUE`/`FORMATTED_VALUE` живого
    API) и `_formulas` (`data_only=False`, для режима `FORMULA`).
    """

    path: Path
    _values: openpyxl.Workbook = field(init=False)
    _formulas: openpyxl.Workbook = field(init=False)

    def __post_init__(self) -> None:
        """Открыть книгу дважды: с кэшированными значениями и с формулами."""
        self._values = openpyxl.load_workbook(self.path, data_only=True)
        self._formulas = openpyxl.load_workbook(self.path, data_only=False)

    def __call__(self, sheet: str, a1: str, value_render_option: str) -> list[Row]:
        """Прочитать один A1-диапазон из открытой книги."""
        parsed = parse_a1_range(a1)
        workbook = self._formulas if value_render_option == _FORMULA else self._values
        worksheet = workbook[parsed.sheet]
        row_start = parsed.row_start if parsed.row_start is not None else 1
        row_end = parsed.row_end if parsed.row_end is not None else row_start
        col_start = column_letter_to_index(parsed.col_start)
        col_end = column_letter_to_index(parsed.col_end)
        rows: list[Row] = []
        for r in range(row_start, row_end + 1):
            row: Row = []
            for c in range(col_start, col_end + 1):
                row.append(_xlsx_cell_value(worksheet, r, c, value_render_option))
            rows.append(row)
        return rows


def _xlsx_cell_value(worksheet: Worksheet, row: int, col: int, value_render_option: str) -> object:
    value = worksheet.cell(row=row, column=col).value
    if value is None:
        return ""
    if value_render_option == _FORMATTED and isinstance(value, datetime | date):
        return (
            value.strftime("%d.%m.%y %H:%M")
            if isinstance(value, datetime)
            else value.strftime("%d.%m.%y")
        )
    return value


def _column_range(col_start: str, col_end: str) -> list[str]:
    start = column_letter_to_index(col_start)
    end = column_letter_to_index(col_end)
    return [column_index_to_letter(i) for i in range(start, end + 1)]


def find_last_data_row(
    reader: RangeReader,
    sheet: str,
    anchor_col: str,
    start_row: int,
    *,
    max_scan_rows: int = _MAX_SCAN_ROWS,
) -> int:
    """Найти номер последней строки блока по «якорной» колонке.

    Колонки в одном физическом листе (`DataBase`) не выровнены по строкам
    между блоками (§I.1: у «Тикетов» 657 строк, у «Юзеров» — 244), поэтому
    у каждого блока свой якорь и своя граница.

    Args:
        reader: Источник данных.
        sheet: Название листа.
        anchor_col: Колонка, заполненная в каждой настоящей строке блока.
        start_row: Первая строка данных (после заголовков).
        max_scan_rows: Сколько строк сканировать за один проход.

    Returns:
        Номер последней непустой строки. `start_row - 1`, если пусто.
    """
    end_row = start_row + max_scan_rows - 1
    values = reader(sheet, a1_range(sheet, anchor_col, start_row, anchor_col, end_row), _FORMATTED)
    last = start_row - 1
    for offset, row in enumerate(values):
        if row and str(row[0]).strip() != "":
            last = start_row + offset
    return last


@dataclass(frozen=True, slots=True)
class SheetBlock:
    """Один экспортируемый прямоугольный блок колонок."""

    key: str
    filename: str
    sheet: str
    col_start: str
    col_end: str
    anchor_col: str
    header: tuple[str, ...]
    #: Колонки этого блока, которые обязаны читаться как `FORMATTED_VALUE`
    #: (сейчас — только колонка `A` «Тикетов», см. §I.3). Остальные колонки
    #: блока читаются как `UNFORMATTED_VALUE`.
    formatted_columns: frozenset[str] = field(default_factory=frozenset)


def _database_block(
    block: DatabaseBlock,
    *,
    anchor_col: str,
    formatted_columns: frozenset[str] = frozenset(),
) -> SheetBlock:
    return SheetBlock(
        key=block.name,
        filename=f"{_slug(block.name)}.csv",
        sheet=DATABASE_SHEET,
        col_start=block.col_start,
        col_end=block.col_end,
        anchor_col=anchor_col,
        header=block.expected_headers,
        formatted_columns=formatted_columns,
    )


def _slug(name: str) -> str:
    return {
        "Тикеты": "tickets",
        "Общая база пользователей": "users",
        "Магазин": "shop",
        "item database": "items",
    }.get(name, name)


#: Блоки СКУПКИ. Якорные колонки: `B` (Ник) для «Тикетов» — заполнена и в
#: строках-заготовках 11/544/545 (§I.10); `J` (Уникальный ник) для «Юзеров»;
#: `U` (Ник) для «Магазина» (пуст сегодня, §I.1); `AA` (id) для каталога.
SKUPKA_BLOCKS: Final[tuple[SheetBlock, ...]] = tuple(
    _database_block(block, anchor_col=anchor, formatted_columns=formatted)
    for block, anchor, formatted in (
        (DATABASE_BLOCKS[0], "B", frozenset({"A"})),  # Тикеты — A читается формой даты
        (DATABASE_BLOCKS[1], "J", frozenset()),  # Общая база пользователей
        (DATABASE_BLOCKS[2], "U", frozenset()),  # Магазин
        (DATABASE_BLOCKS[3], "AA", frozenset()),  # item database
    )
)

#: Верхняя таблица листа «БУСТЫ» (секции Кулинария…Прочее, §I.7),
#: откалибровано `SYNC_LAYOUTS` (`layouts.py`): строки 4-9, 7 колоночных
#: групп по 7 колонок начиная с C.
BOOSTS_TOP_SHEET: Final = "БУСТЫ"
BOOSTS_TOP_ROWS: Final = range(4, 10)
BOOSTS_TOP_NAME_COLS: Final = ("C", "J", "Q", "X", "AE", "AL", "AS")
BOOSTS_TOP_PRICE_COLS: Final = ("D", "K", "R", "Y", "AF", "AM", "AT")

#: Нижняя таблица «БУСТЫ» (§II.3, «Себестоимость»), строки 16-21.
#: Откалибровано по офлайн-снимку живой таблицы (см. docstring модуля):
#: тот же 7-группный раскрой колонок, что и у верхней таблицы (`name`/`price`
#: колонки дословно совпадают — `D16=2818` это себестоимость «Уха» под тем
#: же заголовком «Цена», что в верхней таблице значит цену продажи), только
#: колонки «Кол-во»/«Цена зк» переиспользуются как «Кол-во»/«Выгода».
BOOSTS_BOTTOM_SHEET: Final = "БУСТЫ"
BOOSTS_BOTTOM_ROWS: Final = range(16, 22)
BOOSTS_BOTTOM_NAME_COLS: Final = BOOSTS_TOP_NAME_COLS
BOOSTS_BOTTOM_COST_COLS: Final = BOOSTS_TOP_PRICE_COLS
BOOSTS_BOTTOM_QTY_COLS: Final = ("F", "M", "T", "AA", "AH", "AO", "AV")
BOOSTS_BOTTOM_PROFIT_COLS: Final = ("G", "N", "U", "AB", "AI", "AP", "AW")

#: Итоги листа «БУСТЫ» (под нижней таблицей) — для `test_boost_profit_parity.py`
#: уровня доказательства «сумма/себестоимость/выгода сходятся до рубля» (§VI.2).
BOOSTS_TOTALS_CELLS: Final = {
    "total_revenue": "D22",  # «Общая сумма»
    "total_profit": "D24",  # «Общая выгода»
    "total_cost": "D25",  # «Общая себестоимость»
}

#: Формулы прогрессии, замораживаемые для `test_calculator_matches_frozen_sheet_formulas.py`
#: (§VI.2). Ячейки строки 3 листа `DataBase`.
FORMULA_CELLS: Final = ("F3", "G3", "K3", "L3", "M3", "N3", "O3", "P3", "R3", "S3")


def export_skupka_database(reader: RangeReader, out_dir: Path) -> dict[str, int]:
    """Экспортировать четыре блока `DataBase` в CSV. Возвращает объёмы по ключу."""
    volumes: dict[str, int] = {}
    for block in SKUPKA_BLOCKS:
        last_row = find_last_data_row(reader, block.sheet, block.anchor_col, DATA_START_ROW)
        rows = _read_block_rows(reader, block, DATA_START_ROW, last_row)
        _write_csv(out_dir / block.filename, ("sheet_row", *block.header), rows)
        volumes[block.key] = max(0, last_row - DATA_START_ROW + 1)
        logger.info("%s: %d строк -> %s", block.key, volumes[block.key], block.filename)
    return volumes


def _read_block_rows(
    reader: RangeReader, block: SheetBlock, start_row: int, end_row: int
) -> list[Row]:
    if end_row < start_row:
        return []
    columns = _column_range(block.col_start, block.col_end)
    formatted_cols = [c for c in columns if c in block.formatted_columns]
    plain_cols = [c for c in columns if c not in block.formatted_columns]

    grids: dict[str, list[Row]] = {}
    if formatted_cols:
        grids["formatted"] = reader(
            block.sheet,
            a1_range(block.sheet, formatted_cols[0], start_row, formatted_cols[-1], end_row),
            _FORMATTED,
        )
    if plain_cols:
        grids["plain"] = reader(
            block.sheet,
            a1_range(block.sheet, plain_cols[0], start_row, plain_cols[-1], end_row),
            _UNFORMATTED,
        )

    rows: list[Row] = []
    for offset in range(end_row - start_row + 1):
        sheet_row = start_row + offset
        by_col: dict[str, object] = {}
        if formatted_cols:
            formatted_row = grids["formatted"][offset] if offset < len(grids["formatted"]) else []
            for i, col in enumerate(formatted_cols):
                by_col[col] = formatted_row[i] if i < len(formatted_row) else ""
        if plain_cols:
            plain_row = grids["plain"][offset] if offset < len(grids["plain"]) else []
            for i, col in enumerate(plain_cols):
                by_col[col] = plain_row[i] if i < len(plain_row) else ""
        rows.append([sheet_row, *(by_col[c] for c in columns)])
    return rows


def export_boosts(reader: RangeReader, out_dir: Path) -> int:
    """Экспортировать верхнюю и нижнюю таблицы листа «БУСТЫ» + итоги (§II.3)."""
    top_header = ("sheet_row", "group", "name", "price")
    top_rows: list[Row] = []
    for row_num in BOOSTS_TOP_ROWS:
        for group, (name_col, price_col) in enumerate(
            zip(BOOSTS_TOP_NAME_COLS, BOOSTS_TOP_PRICE_COLS, strict=True), start=1
        ):
            name = reader(
                BOOSTS_TOP_SHEET, a1_range(BOOSTS_TOP_SHEET, name_col, row_num), _UNFORMATTED
            )
            price = reader(
                BOOSTS_TOP_SHEET, a1_range(BOOSTS_TOP_SHEET, price_col, row_num), _UNFORMATTED
            )
            top_rows.append([row_num, group, _cell(name), _cell(price)])
    _write_csv(out_dir / "boosts_top.csv", top_header, top_rows)

    bottom_header = ("sheet_row", "group", "name", "cost", "qty", "profit")
    bottom_rows: list[Row] = []
    for row_num in BOOSTS_BOTTOM_ROWS:
        for group, (name_col, cost_col, qty_col, profit_col) in enumerate(
            zip(
                BOOSTS_BOTTOM_NAME_COLS,
                BOOSTS_BOTTOM_COST_COLS,
                BOOSTS_BOTTOM_QTY_COLS,
                BOOSTS_BOTTOM_PROFIT_COLS,
                strict=True,
            ),
            start=1,
        ):
            name = reader(
                BOOSTS_BOTTOM_SHEET, a1_range(BOOSTS_BOTTOM_SHEET, name_col, row_num), _UNFORMATTED
            )
            cost = reader(
                BOOSTS_BOTTOM_SHEET, a1_range(BOOSTS_BOTTOM_SHEET, cost_col, row_num), _UNFORMATTED
            )
            qty = reader(
                BOOSTS_BOTTOM_SHEET, a1_range(BOOSTS_BOTTOM_SHEET, qty_col, row_num), _UNFORMATTED
            )
            profit = reader(
                BOOSTS_BOTTOM_SHEET,
                a1_range(BOOSTS_BOTTOM_SHEET, profit_col, row_num),
                _UNFORMATTED,
            )
            bottom_rows.append(
                [row_num, group, _cell(name), _cell(cost), _cell(qty), _cell(profit)]
            )
    _write_csv(out_dir / "boosts_bottom.csv", bottom_header, bottom_rows)

    totals = {
        key: _cell(
            reader(
                BOOSTS_BOTTOM_SHEET,
                a1_range(BOOSTS_BOTTOM_SHEET, cell[0], int(cell[1:])),
                _UNFORMATTED,
            )
        )
        for key, cell in BOOSTS_TOTALS_CELLS.items()
    }
    (out_dir / "boosts_totals.json").write_text(
        json.dumps(totals, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8"
    )

    return len(top_rows) + len(bottom_rows)


def _cell(values: list[Row]) -> object:
    if not values or not values[0]:
        return ""
    return values[0][0]


def export_formulas(reader: RangeReader, out_dir: Path) -> None:
    """Заморозить формулы прогрессии строки 3 в `formulas.json` (§VI.2)."""
    formulas: dict[str, str] = {}
    for cell in FORMULA_CELLS:
        values = reader(DATABASE_SHEET, a1_range(DATABASE_SHEET, cell[0], int(cell[1:])), _FORMULA)
        formulas[cell] = str(_cell(values))
    (out_dir / "formulas.json").write_text(
        json.dumps(formulas, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8"
    )


#: --- Убежка (§II) --- Откалибровано по офлайн-снимку живой таблицы
#: (см. docstring модуля): 8 листов профессий + Инструкция/Цены/Тех. лист/
#: Прокачка навыков, все имена листов дословно как в живой книге, включая
#: «Сырье и материалы» (без «ё» — в плане §II написано «Сырьё», в самой
#: таблице «ё» нет).
SHELTER_INSTRUCTION_SHEET: Final = "Инструкция"

SHELTER_PRICES_SHEET: Final = "Цены"
#: 9 блоков с шагом 4 колонки, начиная с A (§II.1): «Основные компоненты» +
#: 8 профессий. Ширина каждого блока — 4 колонки (Предмет/значение1/Моя
#: цена/Итог).
SHELTER_PRICES_BLOCK_STARTS: Final[tuple[str, ...]] = tuple(
    column_index_to_letter(1 + 4 * k) for k in range(9)
)

SHELTER_SKILL_SHEET: Final = "Прокачка навыков"
#: 8 блоков (по одному на профессию, без «Основных компонентов»), ширина 3
#: колонки (Предмет/Цена продажи 1 ед./Цена 1 опыта).
SHELTER_SKILL_BLOCK_STARTS: Final[tuple[str, ...]] = tuple(
    column_index_to_letter(1 + 4 * k) for k in range(8)
)

SHELTER_TECH_SHEET: Final = "Тех. лист"

#: (ключ профессии, название листа) — ключи совпадают с `professions.key`
#: из миграции `0006` (§IV.3).
PROFESSION_SHEETS: Final[tuple[tuple[str, str], ...]] = (
    ("ammo", "Боеприпасы"),
    ("pyro", "Пиротехника"),
    ("armor", "Защитное снаряжение"),
    ("engineering", "Инженерия"),
    ("cooking", "Кулинария"),
    ("moonshine", "Самогоноварение"),
    ("medicine", "Медицина"),
    ("materials", "Сырье и материалы"),
)
#: Маркеры конца блока-рецепта в колонке A/F профессиональных листов (§II.1).
_RECIPE_UNITS_MARKER: Final = "Ед. за крафт"
_RECIPE_COST_MARKER: Final = "Себестоимость"
_RECIPE_HEADER_MARKER: Final = "Предмет"


def _read_table_block(
    reader: RangeReader,
    sheet: str,
    col_start: str,
    data_start_row: int,
    width: int,
) -> list[Row]:
    """Прочитать один прямоугольный блок постоянной ширины (Цены/Тех. лист/…).

    Первая колонка блока — якорь для поиска настоящей границы данных
    (см. `find_last_data_row`); в каждой строке этих блоков она заполнена.
    """
    last_row = find_last_data_row(reader, sheet, col_start, data_start_row)
    if last_row < data_start_row:
        return []
    col_end = column_index_to_letter(column_letter_to_index(col_start) + width - 1)
    values = reader(
        sheet, a1_range(sheet, col_start, data_start_row, col_end, last_row), _UNFORMATTED
    )
    rows: list[Row] = []
    for offset in range(last_row - data_start_row + 1):
        sheet_row = data_start_row + offset
        raw = values[offset] if offset < len(values) else []
        padded = [raw[i] if i < len(raw) else "" for i in range(width)]
        rows.append([sheet_row, *padded])
    return rows


def export_shelter_settings(reader: RangeReader, out_dir: Path) -> int:
    """Экспортировать `Инструкция` (настройки расчёта, §II.1) в CSV."""
    values = reader(
        SHELTER_INSTRUCTION_SHEET,
        a1_range(SHELTER_INSTRUCTION_SHEET, "C", 1, "D", 11),
        _UNFORMATTED,
    )
    rows: list[Row] = []
    for offset, row in enumerate(values):
        key = row[0] if len(row) > 0 else ""
        value = row[1] if len(row) > 1 else ""
        if key in ("", None):
            continue
        rows.append([1 + offset, key, value])
    _write_csv(out_dir / "shelter_settings.csv", ("sheet_row", "key", "value"), rows)
    return len(rows)


def export_shelter_prices(reader: RangeReader, out_dir: Path) -> int:
    """Экспортировать `Цены` — 9 блоков по 4 колонки (§II.1) в один CSV."""
    header = ("block", "sheet_row", "name", "value1", "my_price", "total")
    rows: list[Row] = []
    for block_idx, col in enumerate(SHELTER_PRICES_BLOCK_STARTS):
        for block_row in _read_table_block(reader, SHELTER_PRICES_SHEET, col, 3, 4):
            rows.append([block_idx, *block_row])
    _write_csv(out_dir / "shelter_prices.csv", header, rows)
    return len(rows)


def export_shelter_skill(reader: RangeReader, out_dir: Path) -> int:
    """Экспортировать `Прокачка навыков` — 8 блоков по 3 колонки в один CSV."""
    header = ("block", "sheet_row", "name", "sale_price", "exp_cost")
    rows: list[Row] = []
    for block_idx, col in enumerate(SHELTER_SKILL_BLOCK_STARTS):
        for block_row in _read_table_block(reader, SHELTER_SKILL_SHEET, col, 3, 3):
            rows.append([block_idx, *block_row])
    _write_csv(out_dir / "shelter_skill.csv", header, rows)
    return len(rows)


def export_shelter_tech(reader: RangeReader, out_dir: Path) -> int:
    """Экспортировать три блока `Тех. лист` (§II.1) в три CSV."""
    vendor = _read_table_block(reader, SHELTER_TECH_SHEET, "A", 2, 2)
    _write_csv(out_dir / "shelter_tech_vendor.csv", ("sheet_row", "name", "vendor_price"), vendor)

    yields = _read_table_block(reader, SHELTER_TECH_SHEET, "D", 2, 7)
    _write_csv(
        out_dir / "shelter_tech_yields.csv",
        ("sheet_row", "name", "sale_price", "lvl1", "lvl2", "lvl3", "lvl4", "lvl5"),
        yields,
    )

    costs = _read_table_block(reader, SHELTER_TECH_SHEET, "L", 2, 12)
    _write_csv(
        out_dir / "shelter_tech_costs.csv",
        (
            "sheet_row",
            "name",
            "link1",
            "link2",
            "link3",
            "link4",
            "link5",
            "min_cost",
            "my_price",
            "final_price",
            "min_cost_ref",
            "my_price_ref",
            "final_ref",
        ),
        costs,
    )
    return len(vendor) + len(yields) + len(costs)


@dataclass(frozen=True, slots=True)
class RecipeRaw:
    """Один блок-рецепт, разобранный из колонки профессионального листа."""

    output: str
    source_row: int
    units_per_craft: object
    cost: object
    ingredients: tuple[tuple[str, object, object, int], ...]


def _parse_recipe_column(rows: list[Row], start_row: int) -> list[RecipeRaw]:
    """Разобрать один столбец-раскрой профлиста на блоки-рецепты (§II.1).

    Каждый блок: строка-заголовок с именем крафта (только колонка A/F
    заполнена), строка «Предмет|Количество|Цена 1 ед.|…», N строк
    ингредиентов (включая псевдо-ингредиент «Энергия»), строка
    «Ед. за крафт», строка «Себестоимость» — и сразу следующий блок, без
    гарантированной пустой строки-разделителя.
    """
    recipes: list[RecipeRaw] = []
    i, n = 0, len(rows)
    while i < n:
        row = rows[i]
        name = "" if not row or row[0] in (None, "") else str(row[0]).strip()
        if name == "":
            i += 1
            continue
        qty_val = row[1] if len(row) > 1 else ""
        price_val = row[2] if len(row) > 2 else ""
        if qty_val not in ("", None) or price_val not in ("", None):
            # Не заголовок блока (защитно — не должно происходить в реальных данных).
            i += 1
            continue
        source_row = start_row + i
        i += 1
        if i < n and rows[i] and str(rows[i][0]).strip() == _RECIPE_HEADER_MARKER:
            i += 1
        ingredients: list[tuple[str, object, object, int]] = []
        units_per_craft: object = ""
        cost: object = ""
        position = 0
        while i < n:
            ingredient_row = rows[i]
            cell0 = (
                ""
                if not ingredient_row or ingredient_row[0] in (None, "")
                else str(ingredient_row[0]).strip()
            )
            if cell0 == _RECIPE_UNITS_MARKER:
                units_per_craft = ingredient_row[1] if len(ingredient_row) > 1 else ""
                i += 1
                continue
            if cell0 == _RECIPE_COST_MARKER:
                cost = ingredient_row[1] if len(ingredient_row) > 1 else ""
                i += 1
                break
            if cell0 == "":
                i += 1
                continue
            position += 1
            ing_qty = ingredient_row[1] if len(ingredient_row) > 1 else ""
            ing_price = ingredient_row[2] if len(ingredient_row) > 2 else ""
            ingredients.append((cell0, ing_qty, ing_price, position))
            i += 1
        recipes.append(
            RecipeRaw(
                output=name,
                source_row=source_row,
                units_per_craft=units_per_craft,
                cost=cost,
                ingredients=tuple(ingredients),
            )
        )
    return recipes


def export_shelter_recipes(reader: RangeReader, out_dir: Path) -> int:
    """Экспортировать все рецепты 8 профессиональных листов в один CSV (§II.1-II.2)."""
    header = (
        "profession",
        "output",
        "recipe_seq",
        "source_row",
        "position",
        "ingredient",
        "qty",
        "unit_price",
        "units_per_craft",
        "cost",
    )
    rows: list[Row] = []
    recipe_count = 0
    for key, sheet in PROFESSION_SHEETS:
        left_last = find_last_data_row(reader, sheet, "A", 1)
        right_last = find_last_data_row(reader, sheet, "F", 1)
        left = (
            reader(sheet, a1_range(sheet, "A", 1, "D", left_last), _UNFORMATTED)
            if left_last >= 1
            else []
        )
        right = (
            reader(sheet, a1_range(sheet, "F", 1, "I", right_last), _UNFORMATTED)
            if right_last >= 1
            else []
        )
        recipes = _parse_recipe_column(left, 1) + _parse_recipe_column(right, 1)
        seq_by_output: dict[str, int] = {}
        for recipe in recipes:
            recipe_count += 1
            seq_by_output[recipe.output] = seq_by_output.get(recipe.output, 0) + 1
            seq = seq_by_output[recipe.output]
            if not recipe.ingredients:
                rows.append(
                    [
                        key,
                        recipe.output,
                        seq,
                        recipe.source_row,
                        0,
                        "",
                        "",
                        "",
                        recipe.units_per_craft,
                        recipe.cost,
                    ]
                )
            for ingredient, qty, price, position in recipe.ingredients:
                rows.append(
                    [
                        key,
                        recipe.output,
                        seq,
                        recipe.source_row,
                        position,
                        ingredient,
                        qty,
                        price,
                        recipe.units_per_craft,
                        recipe.cost,
                    ]
                )
    _write_csv(out_dir / "shelter_recipes.csv", header, rows)
    return recipe_count


def export_shelter(reader: RangeReader, out_dir: Path) -> dict[str, int]:
    """Полный экспорт убежки (§II): настройки, цены, прокачка, рецепты."""
    return {
        "shelter_settings": export_shelter_settings(reader, out_dir),
        "shelter_prices": export_shelter_prices(reader, out_dir),
        "shelter_skill": export_shelter_skill(reader, out_dir),
        "shelter_tech": export_shelter_tech(reader, out_dir),
        "shelter_recipes": export_shelter_recipes(reader, out_dir),
    }


def _write_csv(path: Path, header: Sequence[object], rows: Sequence[Row]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for row in rows:
            writer.writerow(["" if cell is None else cell for cell in row])


def build_client(credentials_path: Path) -> gspread.Client:
    """Собрать read-only `gspread`-клиент из service-account credentials."""
    credentials = Credentials.from_service_account_file(  # type: ignore[no-untyped-call]
        str(credentials_path), scopes=_SCOPES
    )
    return gspread.authorize(credentials)


def _skupka_reader(
    *, credentials_path: Path | None, skupka_spreadsheet_id: str | None, skupka_xlsx: Path | None
) -> RangeReader:
    if skupka_xlsx is not None:
        return XlsxRangeReader(skupka_xlsx)
    if credentials_path is None or skupka_spreadsheet_id is None:
        raise ValueError("нужны либо --skupka-xlsx, либо --credentials + --skupka-id")
    client = build_client(credentials_path)
    return GspreadRangeReader(client.open_by_key(skupka_spreadsheet_id))


def _shelter_reader(
    *, credentials_path: Path | None, shelter_spreadsheet_id: str | None, shelter_xlsx: Path | None
) -> RangeReader | None:
    if shelter_xlsx is not None:
        return XlsxRangeReader(shelter_xlsx)
    if shelter_spreadsheet_id is not None and credentials_path is not None:
        client = build_client(credentials_path)
        return GspreadRangeReader(client.open_by_key(shelter_spreadsheet_id))
    return None


def run(
    *,
    credentials_path: Path | None = None,
    skupka_spreadsheet_id: str | None = None,
    skupka_xlsx: Path | None = None,
    shelter_spreadsheet_id: str | None = None,
    shelter_xlsx: Path | None = None,
    out_dir: Path,
) -> dict[str, int]:
    """Полный экспорт: СКУПКА обязательна (живая таблица или `.xlsx`), убежка — опциональна.

    Источник каждой книги — либо живая таблица (`credentials_path` +
    `*_spreadsheet_id`), либо локальный `.xlsx`-снимок (`*_xlsx`). `.xlsx`
    имеет приоритет, если передано и то, и другое.
    """
    skupka_reader = _skupka_reader(
        credentials_path=credentials_path,
        skupka_spreadsheet_id=skupka_spreadsheet_id,
        skupka_xlsx=skupka_xlsx,
    )
    volumes = export_skupka_database(skupka_reader, out_dir)
    volumes["boosts"] = export_boosts(skupka_reader, out_dir)
    export_formulas(skupka_reader, out_dir)

    shelter_reader = _shelter_reader(
        credentials_path=credentials_path,
        shelter_spreadsheet_id=shelter_spreadsheet_id,
        shelter_xlsx=shelter_xlsx,
    )
    if shelter_reader is None:
        logger.warning(
            "Ни --shelter-id, ни --shelter-xlsx не переданы — убежка (§II) не экспортирована."
        )
    else:
        volumes.update(export_shelter(shelter_reader, out_dir))

    return volumes


def main(argv: Sequence[str] | None = None) -> None:
    """CLI-точка входа: разобрать аргументы и запустить полный экспорт."""
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--credentials", type=Path, default=None, help="Путь к service_account.json"
    )
    parser.add_argument("--skupka-id", default=None, help="spreadsheet_id книги СКУПКА")
    parser.add_argument(
        "--skupka-xlsx", type=Path, default=None, help="Локальный .xlsx-снимок СКУПКИ"
    )
    parser.add_argument("--shelter-id", default=None, help="spreadsheet_id книги убежка")
    parser.add_argument(
        "--shelter-xlsx", type=Path, default=None, help="Локальный .xlsx-снимок убежки"
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Каталог снимка (по умолчанию tests/fixtures/sheet_snapshot_<today>/)",
    )
    args = parser.parse_args(argv)

    out_dir = args.out or Path("tests/fixtures") / f"sheet_snapshot_{datetime.now(UTC):%Y-%m-%d}"
    volumes = run(
        credentials_path=args.credentials,
        skupka_spreadsheet_id=args.skupka_id,
        skupka_xlsx=args.skupka_xlsx,
        shelter_spreadsheet_id=args.shelter_id,
        shelter_xlsx=args.shelter_xlsx,
        out_dir=out_dir,
    )
    logger.info("Готово: %s", out_dir)
    logger.info("Объёмы: %s", volumes)


if __name__ == "__main__":
    main()
