"""One-off: extract poster icons/logo/layout from `СКУПКА.xlsx` (Э11).

Часть IX assumed a ready-made background PNG with icons/frames already
drawn, calibrated with a coordinate map. There is no such thing — the
owner confirmed (2026-08-15) the three "screenshot" sheets («Скрин скуп»,
«Скрин бусты», «Скрин скуп бустов») are just the live Google Sheet,
manually screenshotted and posted to Discord. What the workbook *does*
have is everything needed to regenerate an equivalent poster
programmatically: item icons, their reading order, and (for «Скрин
бусты» only) section headers.

Icons are inserted as `oneCellAnchor` drawings, not the classic
`twoCellAnchor` — `openpyxl==3.1.5`'s `worksheet._images` doesn't surface
these, so this script parses `xl/drawings/drawingN.xml` + its `.rels`
directly instead. An anchor's `(col, row)` maps to the adjacent name cell
at `openpyxl(row=row+2, column=col+3)` and price cell at
`(row=row+2, column=col+4)` (1-indexed) — verified by hand against all 24
anchors on «Скрин скуп бустов» before writing this. An anchor with no
text at that name/price pair is decorative (the «Клондайк Шёпота» logo,
reused across all three sheets), not an item.

Output (committed to git, unlike `СКУПКА.xlsx` itself):
- `src/stalbot/assets/posters/icons/<item_name_norm>.png` — deduplicated
  by content hash (the same item's icon often appears on multiple sheets).
- `src/stalbot/assets/posters/logo.png`
- `src/stalbot/assets/posters/layout_<kind>.json` — reading-order list of
  `{name, name_norm, icon}`, grouped into named sections only where the
  sheet actually has them (only «Скрин бусты» does). No prices, no pixel
  coordinates — `PosterService` fills in live prices and the renderer
  computes the grid at draw time.

Lives in `scripts/`, not `src/` — one-shot data extraction, not part of
the coverage denominator (sqlite_migration.md §XI).
"""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import logging
import posixpath
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image

logger = logging.getLogger(__name__)

_XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS = {"xdr": _XDR_NS, "a": _A_NS}

#: Header fill color for section rows on «Скрин бусты» (found by inspection).
_SECTION_FILL = "FF9FC5E8"

#: (sheet name, has sections)
_TARGET_SHEETS: tuple[tuple[str, bool], ...] = (
    ("Скрин скуп", False),
    ("Скрин бусты", True),
    ("Скрин скуп бустов", False),
)


#: Characters `< > : " / \ | ? *` are invalid in a Windows filename — several
#: real item names contain `"` (e.g. `Граната "Кустарник-1"`).
_UNSAFE_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*]')


def _safe_filename(name_norm: str) -> str:
    return _UNSAFE_FILENAME_CHARS.sub("_", name_norm)


def normalize_item_name(name: str) -> str:
    """Collapse whitespace and lower-case an item name for lookup/uniqueness.

    Duplicated from `infrastructure.cache.repositories.items` on purpose —
    this script must not import from `src/stalbot` at module scope in a way
    that would require the whole package to be installed just to run it
    standalone, and the function is one line.
    """
    return " ".join(name.split()).lower()


@dataclass(frozen=True, slots=True)
class _Anchor:
    col: int
    row: int
    media_path: str  # relative to xl/, e.g. "media/image1.png"


def _sheet_name_to_xml(zf: zipfile.ZipFile) -> dict[str, str]:
    workbook_xml = zf.read("xl/workbook.xml")
    rels_xml = zf.read("xl/_rels/workbook.xml.rels")
    # S314: `СКУПКА.xlsx` is the owner's own export, passed by --xlsx-path
    # on the local machine running this one-off script — not attacker
    # data from an untrusted network source.
    wb_root = ET.fromstring(workbook_xml)  # noqa: S314
    rels_root = ET.fromstring(rels_xml)  # noqa: S314

    rid_to_target: dict[str, str] = {}
    for rel in rels_root:
        rid, target = rel.get("Id"), rel.get("Target")
        assert rid is not None and target is not None  # noqa: S101 - always set in a valid .rels
        rid_to_target[rid] = target

    name_to_rid: dict[str, str] = {}
    for sheet in wb_root.iter():
        if sheet.tag.endswith("}sheet"):
            name, rid = sheet.get("name"), sheet.get(f"{{{_R_NS}}}id")
            assert name is not None and rid is not None  # noqa: S101 - always set for a real sheet
            name_to_rid[name] = rid

    return {name: f"xl/{rid_to_target[rid]}" for name, rid in name_to_rid.items()}


def _load_rels(zf: zipfile.ZipFile, rels_path: str) -> dict[str, str]:
    root = ET.fromstring(zf.read(rels_path))  # noqa: S314 - see _sheet_name_to_xml
    rels: dict[str, str] = {}
    for rel in root:
        rid, target = rel.get("Id"), rel.get("Target")
        assert rid is not None and target is not None  # noqa: S101 - always set in a valid .rels
        rels[rid] = target
    return rels


def _drawing_anchors(zf: zipfile.ZipFile, sheet_xml_path: str) -> list[_Anchor]:
    sheet_dir = sheet_xml_path.rsplit("/", 1)[0]
    sheet_rels_path = f"{sheet_dir}/_rels/{sheet_xml_path.rsplit('/', 1)[1]}.rels"
    if sheet_rels_path not in zf.namelist():
        return []
    sheet_rels = _load_rels(zf, sheet_rels_path)
    drawing_target = next((t for t in sheet_rels.values() if "drawing" in t), None)
    if drawing_target is None:
        return []
    drawing_path = posixpath.normpath(f"{sheet_dir}/{drawing_target}")
    drawing_dir = drawing_path.rsplit("/", 1)[0]
    drawing_rels_path = f"{drawing_dir}/_rels/{drawing_path.rsplit('/', 1)[1]}.rels"
    drawing_rels = _load_rels(zf, drawing_rels_path)

    root = ET.fromstring(zf.read(drawing_path))  # noqa: S314 - see _sheet_name_to_xml
    anchors: list[_Anchor] = []
    for anchor in root.findall("xdr:oneCellAnchor", _NS):
        frm = anchor.find("xdr:from", _NS)
        assert frm is not None  # noqa: S101 - every oneCellAnchor has a <from>
        col_el, row_el = frm.find("xdr:col", _NS), frm.find("xdr:row", _NS)
        assert col_el is not None and col_el.text is not None  # noqa: S101
        assert row_el is not None and row_el.text is not None  # noqa: S101
        col, row = int(col_el.text), int(row_el.text)
        blip = anchor.find(".//a:blip", _NS)
        if blip is None:
            continue
        embed = blip.get(f"{{{_R_NS}}}embed")
        target = drawing_rels.get(embed) if embed is not None else None
        if target is None:
            continue
        # `target` is relative to `drawing_dir` (e.g. "../media/x.png") —
        # normalize with posixpath, not `os.path`/`Path`, since zip entry
        # names always use "/" regardless of the host OS.
        media_path = posixpath.normpath(f"{drawing_dir}/{target}")
        anchors.append(_Anchor(col=col, row=row, media_path=media_path))
    return anchors


@dataclass(frozen=True, slots=True)
class _ExtractedItem:
    name: str
    name_norm: str
    icon_bytes: bytes


@dataclass(frozen=True, slots=True)
class _ExtractedSheet:
    logo_candidates: list[bytes]
    """Anchors with no adjacent name/price text — decorative, not items.

    Not necessarily *the* logo: a sheet can have more than one orphaned
    icon with no text next to it for unrelated reasons. `run()` picks
    whichever image recurs across the most sheets as the real logo — it's
    the one thing intentionally duplicated across all three.
    """
    sections: list[tuple[str | None, list[_ExtractedItem]]]


def _extract_sheet(
    zf: zipfile.ZipFile,
    ws: Worksheet,
    sheet_xml_path: str,
    *,
    has_sections: bool,
) -> _ExtractedSheet:
    anchors = _drawing_anchors(zf, sheet_xml_path)

    # Section headers sit several *side by side* in one row (e.g. row 3:
    # "Кулинария" at column B, "Самогоноварение" at F, "Медицина" at J) —
    # keying only by row (as an earlier version of this script did) makes
    # later columns silently overwrite earlier ones in the same row,
    # merging every section before the next header row into one. Keyed by
    # (row, col) instead, resolved by nearest-header-row-above, then
    # nearest-header-column-at-or-left-of the item within that row.
    section_headers: list[tuple[int, int, str]] = []
    if has_sections:
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.value is None:
                    continue
                fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
                if isinstance(fill, str) and fill == _SECTION_FILL and cell.column > 1:
                    section_headers.append((cell.row, cell.column, str(cell.value)))

    def _section_for(row_1indexed: int, col_1indexed: int) -> str | None:
        if not section_headers:
            return None
        header_rows = sorted({r for r, _c, _n in section_headers if r < row_1indexed})
        if not header_rows:
            return None
        block_row = header_rows[-1]
        candidates = [(c, n) for r, c, n in section_headers if r == block_row and c <= col_1indexed]
        if not candidates:
            return None
        return max(candidates, key=lambda pair: pair[0])[1]

    items_by_section: dict[str | None, list[_ExtractedItem]] = {}
    order: list[str | None] = []
    logo_candidates: list[bytes] = []

    for anchor in sorted(anchors, key=lambda a: (a.col, a.row)):
        name_cell = ws.cell(row=anchor.row + 2, column=anchor.col + 3)
        price_cell = ws.cell(row=anchor.row + 2, column=anchor.col + 4)
        image_bytes = zf.read(anchor.media_path)

        if name_cell.value is None and price_cell.value is None:
            logo_candidates.append(image_bytes)
            continue

        name = str(name_cell.value).strip()
        if not name:
            continue
        section = _section_for(anchor.row + 2, anchor.col + 3) if has_sections else None
        if section not in items_by_section:
            items_by_section[section] = []
            order.append(section)
        items_by_section[section].append(
            _ExtractedItem(name=name, name_norm=normalize_item_name(name), icon_bytes=image_bytes)
        )

    sections = [(s, items_by_section[s]) for s in order]
    return _ExtractedSheet(logo_candidates=logo_candidates, sections=sections)


_KIND_BY_SHEET = {
    "Скрин скуп": "resources",
    "Скрин бусты": "boosts",
    "Скрин скуп бустов": "boost_purchases",
}


def run(xlsx_path: Path, assets_dir: Path) -> tuple[dict[str, int], list[str]]:
    """Extract icons/logo/layout for all three poster sheets.

    Args:
        xlsx_path: Path to `СКУПКА.xlsx`.
        assets_dir: `src/stalbot/assets/posters` (created if missing).

    Returns:
        `(kind -> item count, item names whose icon collided with an
        already-used name under different bytes)` — the caller should
        surface the second list for the owner to fix at the source; this
        script doesn't decide which icon is "correct" for those names.
    """
    icons_dir = assets_dir / "icons"
    icons_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    counts: dict[str, int] = {}

    # Content-hash -> already-written icon filename, so the same item's
    # icon (reused across sheets) isn't written twice under two names.
    icon_hash_to_name: dict[str, str] = {}
    # name_norm -> filenames already claimed for it. A handful of real items
    # (found by inspection: гейзер/мякоть лимонника/etc.) have *different*
    # icon bytes under the same name — a copy-paste mistake in the source
    # sheet, not something this script should silently paper over by
    # overwriting one icon file with the other. Suffix the filename instead.
    used_names: dict[str, int] = {}
    name_conflicts: list[str] = []

    with zipfile.ZipFile(xlsx_path) as zf:
        sheet_xml_by_name = _sheet_name_to_xml(zf)

        extracted_by_sheet = {
            sheet_name: _extract_sheet(
                zf, wb[sheet_name], sheet_xml_by_name[sheet_name], has_sections=has_sections
            )
            for sheet_name, has_sections in _TARGET_SHEETS
        }

        # The real logo is the one image reused across the most sheets —
        # a genuinely orphaned icon (no text next to it, but not the logo)
        # only ever shows up on one sheet.
        logo_digest_counts: dict[str, int] = {}
        logo_digest_bytes: dict[str, bytes] = {}
        for extracted in extracted_by_sheet.values():
            seen_this_sheet: set[str] = set()
            for candidate in extracted.logo_candidates:
                digest = hashlib.sha256(candidate).hexdigest()
                logo_digest_bytes[digest] = candidate
                if digest not in seen_this_sheet:
                    logo_digest_counts[digest] = logo_digest_counts.get(digest, 0) + 1
                    seen_this_sheet.add(digest)
        if logo_digest_counts:
            best_digest = max(logo_digest_counts, key=lambda d: logo_digest_counts[d])
            logo = Image.open(io.BytesIO(logo_digest_bytes[best_digest])).convert("RGBA")
            logo.save(assets_dir / "logo.png")

        for sheet_name, _has_sections in _TARGET_SHEETS:
            kind = _KIND_BY_SHEET[sheet_name]
            extracted = extracted_by_sheet[sheet_name]

            layout_sections = []
            item_count = 0
            for section_name, items in extracted.sections:
                layout_items = []
                for item in items:
                    digest = hashlib.sha256(item.icon_bytes).hexdigest()[:16]
                    if digest in icon_hash_to_name:
                        icon_filename = icon_hash_to_name[digest]
                    else:
                        safe_name = _safe_filename(item.name_norm)
                        seen = used_names.get(safe_name, 0)
                        if seen:
                            name_conflicts.append(item.name)
                        icon_filename = (
                            f"{safe_name}.png" if seen == 0 else f"{safe_name}_{seen + 1}.png"
                        )
                        used_names[safe_name] = seen + 1
                        icon = Image.open(io.BytesIO(item.icon_bytes)).convert("RGBA")
                        icon.save(icons_dir / icon_filename)
                        icon_hash_to_name[digest] = icon_filename
                    layout_items.append(
                        {
                            "name": item.name,
                            "name_norm": item.name_norm,
                            "icon": f"icons/{icon_filename}",
                        }
                    )
                    item_count += 1
                layout_sections.append({"name": section_name, "items": layout_items})

            (assets_dir / f"layout_{kind}.json").write_text(
                json.dumps({"sections": layout_sections}, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            counts[kind] = item_count

    return counts, name_conflicts


def main(argv: list[str] | None = None) -> None:
    """CLI entry point."""
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx-path", type=Path, default=Path("СКУПКА.xlsx"))
    parser.add_argument("--assets-dir", type=Path, default=Path("src/stalbot/assets/posters"))
    args = parser.parse_args(argv)

    counts, name_conflicts = run(args.xlsx_path, args.assets_dir)
    for kind, count in counts.items():
        logger.info("%s: %d предметов", kind, count)
    if name_conflicts:
        logger.warning(
            "Разные иконки под одним названием (проверьте вручную, что в таблице "
            "не перепутаны картинки): %s",
            ", ".join(name_conflicts),
        )


if __name__ == "__main__":
    main()
